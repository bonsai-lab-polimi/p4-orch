# spanning_tree.py
"""
SpanningTree class
- builds a simple BFS-like spanning tree among switches (nodes not starting with 'h')
- draws Complete Topology and Spanning Tree (SPT) using networkx + matplotlib if available
- edges avoid node icons (trim endpoints) and use curved arcs for parallel edges
- legend formatting:
    * switch-host: s1:p1-h1
    * switch-switch: s1:p1-p2:s2
- exports an Excel report with:
    * 'Figures' sheet embedding topology and spt PNGs and the legend beside them
    * 'Summary' sheet (root, visited/unvisited, paths to images)
    * 'Log' sheet (last N log lines)
- preserves method signatures and updates global TREE from config
"""

from collections import deque
import logging
import os
import math

# Force headless backend for scripts
import matplotlib
matplotlib.use('Agg')

# Optional drawing libs
try:
    import networkx as nx
    import matplotlib.pyplot as plt
    from matplotlib.offsetbox import OffsetImage, AnnotationBbox
    from matplotlib.patches import FancyArrowPatch
    HAS_DRAW = True
except Exception:
    HAS_DRAW = False

# optional PIL for icons
try:
    from PIL import Image
    HAS_PIL = True
except Exception:
    HAS_PIL = False

# optional numeric support
try:
    import numpy as np
    HAS_NUMPY = True
except Exception:
    HAS_NUMPY = False

# Excel/report libs (used only if present)
try:
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.drawing.image import Image as XLImage
    HAS_EXCEL = True
except Exception:
    HAS_EXCEL = False

# global TREE required by project
from config import TREE

logger = logging.getLogger(__name__)


class SpanningTree:
    def __init__(self, topology):
        """
        topology: dict mapping node -> {neighbor: port, ...}
        Root: lexicographically smallest non-host node (name not starting with 'h').
        """
        self.topology = topology

        # pick root as lexicographically smallest non-host switch
        switch_candidates = [n for n in topology.keys() if not str(n).startswith('h')]
        self.root = min(switch_candidates) if switch_candidates else 's1'

        self.spanning_tree = {}

        # outputs & options
        self.log_file = "spanning_tree.log"
        self.topology_image = "topology_complete.png"
        self.spt_image = "spanning_tree.png"
        self.excel_report = "spanning_report.xlsx"
        self.export_excel = True               # toggle Excel export
        self.excel_embed_figures = True        # embed the two PNGs into 'Figures' sheet
        self.excel_log_lines = 500

        # icons (auto-detected from Images/)
        self.host_icon = "Images/pc.png" if os.path.exists("Images/pc.png") else None
        self.switch_icon = "Images/switch.png" if os.path.exists("Images/switch.png") else None
        self.icon_scale_switch = 0.09
        self.icon_scale_host = 0.065

        # layout tuning
        self.full_k = 1.0
        self.spt_k = 1.8
        self.layout_iterations = 300

        # geometry for trimming edges around icons
        self.node_radius_factor = 0.04
        self.switch_radius_multiplier = 1.25
        self.host_radius_multiplier = 0.9

        # curvature for parallel edges
        self.max_rad = 0.42

        # file logger
        self._file_logger = None
        self._setup_file_logger(self.log_file)

    # ---------------- logging ----------------
    def _setup_file_logger(self, filename):
        if self._file_logger is not None:
            for h in list(self._file_logger.handlers):
                self._file_logger.removeHandler(h)
        self._file_logger = logging.getLogger(f"{__name__}.filelogger.{filename}")
        self._file_logger.setLevel(logging.INFO)
        existing_files = [getattr(h, "baseFilename", None) for h in self._file_logger.handlers]
        if filename not in existing_files:
            fh = logging.FileHandler(filename, mode='w')
            fmt = logging.Formatter("%(asctime)s - %(levelname)s - %(message)s")
            fh.setFormatter(fmt)
            self._file_logger.addHandler(fh)

    # ---------------- topology validators/formatters ----------------
    def _validate_topology(self):
        for a, neigh in self.topology.items():
            for b in neigh:
                if b in self.topology and a not in self.topology[b]:
                    msg = f"Link inconsistency: {a} -> {b} but not {b} -> {a}"
                    logger.warning(msg)
                    if self._file_logger:
                        self._file_logger.warning(msg)

    def _format_legend_label(self, a, b):
        """
        Legend formats:
          - switch-host: s1:p1-h1  (switch port only)
          - switch-switch: s1:p1-p2:s2
        Use '?' where port information is missing.
        """
        a_is_host = str(a).startswith('h')
        b_is_host = str(b).startswith('h')

        if a_is_host and not b_is_host:
            pb = self.topology.get(b, {}).get(a)
            return f"{b}:{pb if pb is not None else '?'}-{a}"
        if b_is_host and not a_is_host:
            pa = self.topology.get(a, {}).get(b)
            return f"{a}:{pa if pa is not None else '?'}-{b}"

        pa = self.topology.get(a, {}).get(b)
        pb = self.topology.get(b, {}).get(a)
        return f"{a}:{pa if pa is not None else '?'}-{pb if pb is not None else '?'}:{b}"

    # ---------------- geometry helpers ----------------
    def _compute_node_radius_map(self, pos):
        xs = [p[0] for p in pos.values()]
        ys = [p[1] for p in pos.values()]
        if not xs or not ys:
            return {n: 0.02 for n in pos}
        width = max(xs) - min(xs)
        height = max(ys) - min(ys)
        span = max(width, height) if max(width, height) > 0 else 1.0
        base = span * self.node_radius_factor
        radii = {}
        for n in pos:
            if str(n).startswith('h'):
                radii[n] = base * self.host_radius_multiplier
            else:
                radii[n] = base * self.switch_radius_multiplier
        return radii

    def _trim_point_outside_node(self, p_from, p_to, radius_from):
        x1, y1 = p_from
        x2, y2 = p_to
        dx = x2 - x1
        dy = y2 - y1
        dist = math.hypot(dx, dy)
        if dist == 0:
            return (x1, y1)
        ux, uy = dx / dist, dy / dist
        return (x1 + ux * radius_from, y1 + uy * radius_from)

    def _assign_edge_rads(self, G):
        """
        For edges that appear multiple times between same node pair (parallel edges),
        compute a set of rad values to space them out.
        """
        rads = {}
        groups = {}
        for u, v in G.edges():
            key = frozenset((u, v))
            groups.setdefault(key, []).append((u, v))
        for key, edges in groups.items():
            m = len(edges)
            if m == 1:
                rads[edges[0]] = 0.0
                continue
            max_rad = self.max_rad
            if m % 2 == 1:
                k = (m - 1) // 2
                steps = [(i + 1) * (max_rad / (k + 1)) for i in range(k)]
                vals = [-x for x in reversed(steps)] + [0.0] + steps
            else:
                k = m // 2
                steps = [(i + 1) * (max_rad / k) for i in range(k)]
                vals = [-x for x in reversed(steps)] + steps
            for (edge, rad) in zip(edges, vals):
                rads[edge] = rad
        return rads

    # ---------------- drawing helpers ----------------
    def _draw_icons_per_node(self, ax, pos, nodes, icon_path, zoom):
        if not HAS_PIL or not icon_path or not os.path.exists(icon_path):
            return False
        try:
            img = Image.open(icon_path)
        except Exception as e:
            logger.warning("Failed to open icon %s: %s", icon_path, e)
            if self._file_logger:
                self._file_logger.warning("Failed to open icon %s: %s", icon_path, e)
            return False
        placed = False
        for n in nodes:
            if n not in pos:
                continue
            x, y = pos[n]
            try:
                offset_img = OffsetImage(img, zoom=zoom)
                ab = AnnotationBbox(offset_img, (x, y), frameon=False, pad=0.02, zorder=60)
                ax.add_artist(ab)
                placed = True
            except Exception as e:
                logger.warning("Error placing icon for node %s: %s", n, e)
                if self._file_logger:
                    self._file_logger.warning("Error placing icon for node %s: %s", n, e)
        return placed

    def _draw_edge_avoiding_nodes(self, ax, p1, p2, r1, r2, rad, color, width, zorder=20):
        t1 = self._trim_point_outside_node(p1, p2, r1)
        t2 = self._trim_point_outside_node(p2, p1, r2)
        arrow = FancyArrowPatch(posA=t1, posB=t2,
                                connectionstyle=f"arc3,rad={rad}",
                                arrowstyle='-', linewidth=width, color=color,
                                shrinkA=0, shrinkB=0, zorder=zorder)
        ax.add_patch(arrow)

    # ---------------- draw full topology ----------------
    def _draw_full_topology(self, filename=None):
        filename = filename or self.topology_image
        if not HAS_DRAW:
            msg = "Drawing libs not available (networkx/matplotlib). Skipping full topology image."
            logger.warning(msg)
            if self._file_logger:
                self._file_logger.warning(msg)
            return

        # build graph including hosts
        G = nx.Graph()
        for a, neigh in self.topology.items():
            G.add_node(a, type='host' if str(a).startswith('h') else 'switch')
            for b, _ in neigh.items():
                G.add_node(b, type='host' if str(b).startswith('h') else 'switch')
                if not G.has_edge(a, b):
                    G.add_edge(a, b)

        # legend labels for undirected pairs
        undirected_labels = {}
        for u, v in G.edges():
            key = tuple(sorted((u, v)))
            undirected_labels[key] = self._format_legend_label(u, v)

        # layout
        try:
            pos = nx.spring_layout(G, seed=42, k=self.full_k, iterations=self.layout_iterations)
        except Exception:
            pos = nx.random_layout(G)

        # radii and edge curvatures
        radii = self._compute_node_radius_map(pos)
        edge_rads = self._assign_edge_rads(G)

        # figure + axes
        fig = plt.figure(figsize=(16, 10))
        ax_main = fig.add_axes([0.02, 0.05, 0.72, 0.92])
        ax_legend = fig.add_axes([0.76, 0.05, 0.22, 0.92])
        ax_legend.axis('off')

        # draw edges first (they will be under icons)
        for u, v in G.edges():
            rad = edge_rads.get((u, v), edge_rads.get((v, u), 0.0))
            p1 = pos[u]
            p2 = pos[v]
            r1 = radii.get(u, 0.02)
            r2 = radii.get(v, 0.02)
            self._draw_edge_avoiding_nodes(ax_main, p1, p2, r1, r2, rad, color="#6c6c6c", width=1.6, zorder=10)

        # detect icons
        if not self.switch_icon and os.path.exists("Images/switch.png"):
            self.switch_icon = "Images/switch.png"
        if not self.host_icon and os.path.exists("Images/pc.png"):
            self.host_icon = "Images/pc.png"

        switches = [n for n, d in G.nodes(data=True) if d.get('type') == 'switch']
        hosts = [n for n, d in G.nodes(data=True) if d.get('type') == 'host']

        drew_switch_icons = False
        drew_host_icons = False
        if self.switch_icon and HAS_PIL and switches:
            drew_switch_icons = self._draw_icons_per_node(ax_main, pos, switches, self.switch_icon, self.icon_scale_switch)
        if self.host_icon and HAS_PIL and hosts:
            drew_host_icons = self._draw_icons_per_node(ax_main, pos, hosts, self.host_icon, self.icon_scale_host)

        # fallback nodes if icons missing
        if not drew_switch_icons:
            nx.draw_networkx_nodes(G, pos, nodelist=switches, node_shape='s', node_color="#2c7bb6",
                                   node_size=900, ax=ax_main, zorder=60)
        if not drew_host_icons:
            nx.draw_networkx_nodes(G, pos, nodelist=hosts, node_shape='o', node_color="#de8f05",
                                   node_size=500, ax=ax_main, zorder=60)

        # node labels under icons
        xs = [p[0] for p in pos.values()]
        ys = [p[1] for p in pos.values()]
        minx, maxx = min(xs), max(xs)
        miny, maxy = min(ys), max(ys)
        width = maxx - minx if maxx - minx > 0 else 1.0
        height = maxy - miny if maxy - miny > 0 else 1.0
        span = max(width, height)
        label_offset = span * 0.03

        for n, (x, y) in pos.items():
            ax_main.text(x, y - label_offset, str(n), fontsize=9, ha='center', va='top', zorder=100)

        # ensure axes limits include icons (avoid clipping)
        max_radius = max(radii.values()) if radii else span * 0.03
        pad = max_radius * 1.6 + span * 0.03
        ax_main.set_xlim(minx - pad, maxx + pad)
        ax_main.set_ylim(miny - pad, maxy + pad)
        ax_main.set_aspect('equal')
        ax_main.set_title("Complete Topology", fontsize=14, fontweight='bold')
        ax_main.axis('off')

        # legend area: icons + formatted list (will be also exported to Excel)
        y_cursor = 0.96
        line_height = 0.055

        # icons top
        if self.switch_icon and HAS_PIL:
            try:
                img = Image.open(self.switch_icon)
                offset_img = OffsetImage(img, zoom=0.06)
                ab = AnnotationBbox(offset_img, (0.06, y_cursor), xycoords='axes fraction', frameon=False, pad=0.02)
                ax_legend.add_artist(ab)
                ax_legend.text(0.14, y_cursor, "Switch", transform=ax_legend.transAxes, va='center', fontsize=10)
            except Exception:
                ax_legend.text(0.08, y_cursor, "Switch", transform=ax_legend.transAxes, va='center', fontsize=10)
        else:
            ax_legend.text(0.08, y_cursor, "Switch", transform=ax_legend.transAxes, va='center', fontsize=10)

        y_cursor -= line_height
        if self.host_icon and HAS_PIL:
            try:
                img = Image.open(self.host_icon)
                offset_img = OffsetImage(img, zoom=0.05)
                ab = AnnotationBbox(offset_img, (0.06, y_cursor), xycoords='axes fraction', frameon=False, pad=0.02)
                ax_legend.add_artist(ab)
                ax_legend.text(0.14, y_cursor, "Host", transform=ax_legend.transAxes, va='center', fontsize=10)
            except Exception:
                ax_legend.text(0.08, y_cursor, "Host", transform=ax_legend.transAxes, va='center', fontsize=10)
        else:
            ax_legend.text(0.08, y_cursor, "Host", transform=ax_legend.transAxes, va='center', fontsize=10)

        # separator and list (but guard long lists)
        y_cursor -= (line_height * 0.7)
        ax_legend.hlines(y_cursor, xmin=0.02, xmax=0.96, transform=ax_legend.transAxes, colors="#cccccc", linewidth=1)
        y_cursor -= (line_height * 0.6)

        # sorted undirected keys
        sorted_keys = sorted(undirected_labels.keys())
        fontsize = 9
        max_lines = int((y_cursor - 0.02) / line_height)
        count = 0
        for key in sorted_keys:
            if count >= max_lines:
                remaining = len(sorted_keys) - count
                ax_legend.text(0.02, 0.02, f"... ({remaining} more)", transform=ax_legend.transAxes, fontsize=9)
                break
            lab = undirected_labels[key]
            line_y = y_cursor - count * line_height
            ax_legend.plot([0.02, 0.08], [line_y, line_y], transform=ax_legend.transAxes, color="#6c6c6c", linewidth=3)
            ax_legend.text(0.10, line_y, lab, transform=ax_legend.transAxes, va='center', fontsize=fontsize)
            count += 1

        plt.savefig(filename, dpi=220, bbox_inches='tight', pad_inches=0.06)
        plt.close(fig)

        msg = f"Full topology image saved to {filename}"
        logger.info(msg)
        if self._file_logger:
            self._file_logger.info(msg)

    # ---------------- draw SPT ----------------
    def _draw_spt(self, filename=None):
        filename = filename or self.spt_image
        if not HAS_DRAW:
            msg = "Drawing libs not available (networkx/matplotlib). Skipping SPT image."
            logger.warning(msg)
            if self._file_logger:
                self._file_logger.warning(msg)
            return

        # build switch-only graph from spanning_tree
        G = nx.Graph()
        for a, neigh in self.spanning_tree.items():
            G.add_node(a, type='switch')
            for b in neigh:
                if not G.has_edge(a, b):
                    G.add_edge(a, b)

        if G.number_of_nodes() == 0:
            msg = "Spanning tree is empty; no SPT image generated."
            logger.info(msg)
            if self._file_logger:
                self._file_logger.info(msg)
            return

        try:
            pos = nx.spring_layout(G, seed=42, k=self.spt_k, iterations=self.layout_iterations)
        except Exception:
            pos = nx.random_layout(G)

        radii = self._compute_node_radius_map(pos)
        edge_rads = self._assign_edge_rads(G)

        fig, ax = plt.subplots(figsize=(14, 10))

        # draw SPT edges first (under nodes)
        for (u, v) in G.edges():
            rad = edge_rads.get((u, v), edge_rads.get((v, u), 0.0))
            p1 = pos[u]; p2 = pos[v]
            r1 = radii.get(u, 0.02); r2 = radii.get(v, 0.02)
            self._draw_edge_avoiding_nodes(ax, p1, p2, r1, r2, rad, color="#d7191c", width=3.0, zorder=5)

        # icons or fallback nodes
        if not self.switch_icon and os.path.exists("Images/switch.png"):
            self.switch_icon = "Images/switch.png"
        drew_icons = False
        if self.switch_icon and HAS_PIL:
            drew_icons = self._draw_icons_per_node(ax, pos, list(G.nodes()), self.switch_icon, self.icon_scale_switch)
        if not drew_icons:
            nx.draw_networkx_nodes(G, pos, nodelist=list(G.nodes()), node_shape='s', node_color="#2c7bb6",
                                   node_size=1000, ax=ax, zorder=40)

        # labels under nodes
        xs = [p[0] for p in pos.values()]
        ys = [p[1] for p in pos.values()]
        minx, maxx = min(xs), max(xs)
        miny, maxy = min(ys), max(ys)
        width = maxx - minx if maxx - minx > 0 else 1.0
        height = maxy - miny if maxy - miny > 0 else 1.0
        span = max(width, height)
        label_offset = span * 0.035

        for n, (x, y) in pos.items():
            ax.text(x, y - label_offset, str(n), fontsize=10, ha='center', va='top', zorder=60)

        # root above node (red text)
        if self.root in G.nodes():
            rx, ry = pos[self.root]
            ax.text(rx, ry + label_offset, "root", fontsize=12, fontweight='bold', ha='center', color='red', zorder=60)

        # axis limits to avoid clipping
        max_radius = max(radii.values()) if radii else span * 0.03
        pad = max_radius * 1.6 + span * 0.03
        ax.set_xlim(minx - pad, maxx + pad)
        ax.set_ylim(miny - pad, maxy + pad)
        ax.set_aspect('equal')

        ax.set_title("Spanning Tree (SPT)", fontsize=14, fontweight='bold')
        ax.axis('off')
        plt.savefig(filename, dpi=220, bbox_inches='tight', pad_inches=0.06)
        plt.close(fig)

        msg = f"SPT image saved to {filename}"
        logger.info(msg)
        if self._file_logger:
            self._file_logger.info(msg)

    # ---------------- export Excel with embedded figures ----------------
    def _export_report(self, excel_filename=None, log_last_n=None, embed_images=True):
        """
        Create an Excel workbook with:
         - 'Figures' sheet: embeds topology and spt images and writes legend next to them
         - 'Topology' sheet: tabular list of undirected links with ports
         - 'SPT' sheet: spanning tree edges (switch-switch, ports if available)
         - 'Legend' sheet: formatted legend entries
         - 'Summary' sheet: root, visited/unvisited, image paths
         - 'Log' sheet: last N lines from log
        Embedding images requires openpyxl + pillow; if not available, only sheets are written.
        """
        excel_filename = excel_filename or self.excel_report
        log_last_n = log_last_n or self.excel_log_lines

        # prepare tabular data
        seen = set()
        topo_rows = []
        legend_rows = []
        for a, neigh in self.topology.items():
            for b, pa in neigh.items():
                key = tuple(sorted((a, b)))
                if key in seen:
                    continue
                seen.add(key)
                pb = self.topology.get(b, {}).get(a)
                topo_rows.append({"node_a": a, "node_b": b, "port_a": pa if pa is not None else None, "port_b": pb if pb is not None else None})
                legend_rows.append({"edge": f"{a} <-> {b}", "label": self._format_legend_label(a, b)})

        # spt rows
        spt_seen = set()
        spt_rows = []
        for parent, neigh in self.spanning_tree.items():
            for child, port_p in neigh.items():
                if parent == child:
                    continue
                key = tuple(sorted((parent, child)))
                if key in spt_seen:
                    continue
                spt_seen.add(key)
                port_c = self.spanning_tree.get(child, {}).get(parent)
                spt_rows.append({"switch_a": parent, "switch_b": child, "port_a": port_p if port_p is not None else None, "port_b": port_c if port_c is not None else None})

        # summary
        visited = set([n for n in self.spanning_tree.keys() if self.spanning_tree.get(n)])
        if not visited and self.root in self.topology:
            visited = {self.root}
        all_switches = [n for n in self.topology.keys() if not str(n).startswith('h')]
        unvisited = sorted(list(set(all_switches) - visited))
        summary = {
            "root": self.root,
            "visited_count": len(visited),
            "visited_nodes": ", ".join(sorted(visited)),
            "unvisited_count": len(unvisited),
            "unvisited_nodes": ", ".join(unvisited),
            "topology_image": os.path.abspath(self.topology_image),
            "spt_image": os.path.abspath(self.spt_image),
            "log_file": os.path.abspath(self.log_file)
        }

        # log lines
        log_lines = []
        if os.path.exists(self.log_file):
            try:
                with open(self.log_file, 'r', encoding='utf-8', errors='ignore') as f:
                    all_lines = f.readlines()
                    log_lines = [l.rstrip("\n") for l in all_lines[-log_last_n:]]
            except Exception as e:
                logger.warning("Unable to read log file: %s", e)
                if self._file_logger:
                    self._file_logger.warning("Unable to read log file: %s", e)

        # write sheets with pandas if available; otherwise try fallback minimal CSVs
        if HAS_EXCEL:
            try:
                with pd.ExcelWriter(excel_filename, engine="openpyxl") as writer:
                    pd.DataFrame(topo_rows).to_excel(writer, sheet_name="Topology", index=False)
                    pd.DataFrame(spt_rows).to_excel(writer, sheet_name="SPT", index=False)
                    pd.DataFrame(legend_rows).to_excel(writer, sheet_name="Legend", index=False)
                    pd.DataFrame([summary]).to_excel(writer, sheet_name="Summary", index=False)
                    pd.DataFrame({"log": log_lines}).to_excel(writer, sheet_name="Log", index=False)
                    # create empty 'Figures' sheet; we'll embed images using openpyxl
                    pd.DataFrame({"notes": ["Figures sheet contains embedded PNGs and the legend"]}).to_excel(writer, sheet_name="Figures", index=False)
            except Exception as e:
                logger.exception("Failed to write base Excel: %s", e)
                if self._file_logger:
                    self._file_logger.exception("Failed to write base Excel: %s", e)
                return

            # embed images and write legend in 'Figures'
            if self.excel_embed_figures:
                try:
                    wb = load_workbook(excel_filename)
                    ws = wb["Figures"]
                    # anchors (tweakable)
                    topology_anchor_row = 1
                    spt_anchor_row = 28
                    legend_col = 8  # column H
                    legend_start_row = 2

                    # embed topology image
                    if os.path.exists(self.topology_image):
                        try:
                            img_top = XLImage(self.topology_image)
                            max_w = 900
                            if img_top.width > max_w:
                                ratio = max_w / img_top.width
                                img_top.width = int(img_top.width * ratio)
                                img_top.height = int(img_top.height * ratio)
                            ws.add_image(img_top, f"A{topology_anchor_row}")
                        except Exception as e:
                            logger.warning("Failed to embed topology image: %s", e)
                            if self._file_logger:
                                self._file_logger.warning("Failed to embed topology image: %s", e)
                    else:
                        logger.warning("Topology image not found for embedding: %s", self.topology_image)

                    # embed spt image
                    if os.path.exists(self.spt_image):
                        try:
                            img_spt = XLImage(self.spt_image)
                            max_w = 900
                            if img_spt.width > max_w:
                                ratio = max_w / img_spt.width
                                img_spt.width = int(img_spt.width * ratio)
                                img_spt.height = int(img_spt.height * ratio)
                            ws.add_image(img_spt, f"A{spt_anchor_row}")
                        except Exception as e:
                            logger.warning("Failed to embed SPT image: %s", e)
                            if self._file_logger:
                                self._file_logger.warning("Failed to embed SPT image: %s", e)
                    else:
                        logger.warning("SPT image not found for embedding: %s", self.spt_image)

                    # write legend next to images
                    ws.cell(row=1, column=legend_col, value="Legend (edge → formatted label)")
                    ws.cell(row=1, column=legend_col).font = ws.cell(row=1, column=legend_col).font.copy(bold=True)
                    ws.cell(row=legend_start_row, column=legend_col, value="edge")
                    ws.cell(row=legend_start_row, column=legend_col + 1, value="label")
                    r = legend_start_row + 1
                    for item in legend_rows:
                        ws.cell(row=r, column=legend_col, value=item["edge"])
                        ws.cell(row=r, column=legend_col + 1, value=item["label"])
                        r += 1

                    # small note
                    ws.cell(row=r + 1, column=legend_col, value="Note: formats s1:p1-p2:s2 and s1:p1-h1 used.")
                    wb.save(excel_filename)
                except Exception as e:
                    logger.exception("Failed to embed images/legend into Excel: %s", e)
                    if self._file_logger:
                        self._file_logger.exception("Failed to embed images/legend into Excel: %s", e)
        else:
            # No pandas/openpyxl: write CSV fallbacks
            try:
                base, _ = os.path.splitext(excel_filename)
                topo_csv = base + "_topology.csv"
                spt_csv = base + "_spt.csv"
                legend_csv = base + "_legend.csv"
                summary_csv = base + "_summary.csv"
                log_csv = base + "_log.csv"
                # write simple CSVs using built-in
                import csv
                with open(topo_csv, 'w', newline='', encoding='utf-8') as f:
                    w = csv.DictWriter(f, fieldnames=["node_a", "node_b", "port_a", "port_b"])
                    w.writeheader()
                    w.writerows(topo_rows)
                with open(spt_csv, 'w', newline='', encoding='utf-8') as f:
                    w = csv.DictWriter(f, fieldnames=["switch_a", "switch_b", "port_a", "port_b"])
                    w.writeheader()
                    w.writerows(spt_rows)
                with open(legend_csv, 'w', newline='', encoding='utf-8') as f:
                    w = csv.DictWriter(f, fieldnames=["edge", "label"])
                    w.writeheader()
                    w.writerows(legend_rows)
                with open(summary_csv, 'w', newline='', encoding='utf-8') as f:
                    w = csv.DictWriter(f, fieldnames=list(summary.keys()))
                    w.writeheader()
                    w.writerow(summary)
                with open(log_csv, 'w', newline='', encoding='utf-8') as f:
                    for line in log_lines:
                        f.write(line + "\n")
                msg = f"Excel libs not available. Wrote CSV fallbacks: {topo_csv}, {spt_csv}, {legend_csv}, {summary_csv}, {log_csv}"
                logger.info(msg)
                if self._file_logger:
                    self._file_logger.info(msg)
            except Exception as e:
                logger.exception("Failed to write CSV fallbacks: %s", e)
                if self._file_logger:
                    self._file_logger.exception("Failed to write CSV fallbacks: %s", e)

    # ---------------- summary writer ----------------
    def _write_summary(self, visited_set, unvisited_set):
        root = self.root
        visited_count = len(visited_set)
        unvisited_count = len(unvisited_set)
        visited_list = sorted(list(visited_set))
        unvisited_list = sorted(list(unvisited_set))
        lines = [
            f"Root: {root}",
            f"Visited count: {visited_count}",
            f"Visited nodes: {visited_list}",
            f"Unvisited count: {unvisited_count}",
            f"Unvisited nodes: {unvisited_list}",
            f"Full topology image: {os.path.abspath(self.topology_image)}",
            f"SPT image: {os.path.abspath(self.spt_image)}",
            f"Spanning tree entries: {self.spanning_tree}"
        ]
        if self._file_logger:
            for ln in lines:
                self._file_logger.info(ln)
        else:
            for ln in lines:
                logger.info(ln)

    # ---------------- core algorithm (public) ----------------
    def build_tree(self):
        """
        Build the spanning tree, draw topology and spt images (best-effort), update TREE,
        write summary and optionally export Excel report with embedded figures.
        """
        try:
            self._validate_topology()

            # filter out hosts
            filtered_topology = {
                switch: {neighbor: port for neighbor, port in neighbors.items() if not neighbor.startswith('h')}
                for switch, neighbors in self.topology.items()
            }

            if self.root not in filtered_topology:
                raise KeyError(f"Root switch {self.root} not found in the topology.")

            self.spanning_tree = {switch: {} for switch in filtered_topology}
            visited = set([self.root])
            queue = deque()

            # initialize queue deterministically
            for neighbor, port in sorted(filtered_topology[self.root].items()):
                queue.append((self.root, neighbor, port))

            while queue:
                parent, child, port = queue.popleft()
                if child in visited:
                    continue
                child_ports = filtered_topology.get(child, {})
                reverse_port = child_ports.get(parent)
                if reverse_port is None:
                    msg = f"Missing reverse port for edge {parent}-{child}; skipping."
                    logger.warning(msg)
                    if self._file_logger:
                        self._file_logger.warning(msg)
                    continue

                self.spanning_tree[parent][child] = port
                self.spanning_tree[child][parent] = reverse_port
                visited.add(child)

                for nb, nb_port in sorted(child_ports.items()):
                    if nb not in visited:
                        queue.append((child, nb, nb_port))

            unvisited = set(filtered_topology.keys()) - visited
            if unvisited:
                msg = f"Unreached switches from root {self.root}: {sorted(unvisited)}"
                logger.warning(msg)
                if self._file_logger:
                    self._file_logger.warning(msg)

            # print SPT to stdout
            self.print_tree()

            # update global TREE
            global TREE
            try:
                TREE.clear()
                TREE.update(self.spanning_tree)
            except Exception:
                logger.exception("Failed to update global TREE; reinitializing.")
                TREE = {}
                TREE.update(self.spanning_tree)

            # auto-detect icons again in case they were added during runtime
            if not self.host_icon and os.path.exists("Images/pc.png"):
                self.host_icon = "Images/pc.png"
            if not self.switch_icon and os.path.exists("Images/switch.png"):
                self.switch_icon = "Images/switch.png"

            # generate drawings
            try:
                self._draw_full_topology(self.topology_image)
                self._draw_spt(self.spt_image)
            except Exception:
                logger.exception("Error while drawing topology or SPT (continuing).")
                if self._file_logger:
                    self._file_logger.exception("Error while drawing topology or SPT (continuing).")

            # write summary to log and export report
            self._write_summary(visited, unvisited)
            if self.export_excel:
                self._export_report(self.excel_report, log_last_n=self.excel_log_lines, embed_images=self.excel_embed_figures)

            logger.info("Spanning tree built successfully. Visited %d nodes.", len(visited))
            if self._file_logger:
                self._file_logger.info("Spanning tree built successfully. Visited %d nodes.", len(visited))

        except KeyError as e:
            logger.error("KeyError: %s. Check root/topology.", e)
            if self._file_logger:
                self._file_logger.error("KeyError: %s. Check root/topology.", e)
        except Exception:
            logger.exception("Unexpected error while building spanning tree.")
            if self._file_logger:
                self._file_logger.exception("Unexpected error while building spanning tree.")

    def print_tree(self):
        print("Spanning Tree:")
        for switch, connections in self.spanning_tree.items():
            for neighbor, port in connections.items():
                print(f"{switch} -({port})-> {neighbor}")

    def get_tree(self):
        return self.spanning_tree
